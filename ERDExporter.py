import streamlit as st
import openpyxl
import warnings
from io import BytesIO
import pandas as pd

def process_excel_file(file, is_paid_version):
    """단일 Excel 파일 처리"""
    try:
        # 파일 내용을 BytesIO 객체로 읽기
        bytes_data = BytesIO(file.read())
        
        # 메모리 최적화를 위해 read_only=True 사용
        workbook = openpyxl.load_workbook(bytes_data, data_only=True, read_only=True)
        
        output_text = ""
        filename = file.name
        base_filename = ('Table ' + filename.replace('DataDefine_', '').replace('.xlsx', '')) if 'DataDefine_' in filename else ('Table ' + filename.replace('.xlsx', ''))
        
        # Define 시트 처리
        if 'Define' in workbook.sheetnames:
            # TableDefine 시트 처리
            if 'TableDefine' in workbook.sheetnames:
                sheet2 = workbook['TableDefine']
                a2_value = None
                b2_value = None
                for row in sheet2.iter_rows(min_row=2, max_row=2, values_only=True):
                    if row:
                        a2_value = row[0]
                        b2_value = row[1] if len(row) > 1 else None
                        break
                
                if is_paid_version:
                    if b2_value is not None:
                        output_text += f"{base_filename} [Note: '{a2_value}', headercolor:{b2_value}]\n{{\n"
                    else:
                        output_text += f"{base_filename} [Note: '{a2_value}', headercolor:#000000]\n{{\n"
                else:
                    output_text += f"{base_filename} [Note: '{a2_value}']\n{{\n"
            else:
                output_text += f"{base_filename}\n{{\n"
            
            # Define 시트의 데이터 처리
            sheet = workbook['Define']
            columns = None
            for i, row in enumerate(sheet.iter_rows(values_only=True)):
                if i == 0:  # 헤더 행
                    columns = row
                    continue
                    
                if len(row) >= 11:  # 최소 필요 컬럼 수 확인
                    a_value, b_value, c_value = row[0], row[1], row[2]
                    IsNotNull = row[4]
                    DefaultKey = row[10]
                    
                    if a_value is not None and b_value is not None:
                        c_value = str(c_value).replace('\n', ' ') if c_value else ''
                        DefaultKeyResult = f'"{DefaultKey}"' if DefaultKey else None
                        
                        output_line = (
                            f"{a_value} {a_value if b_value == 'enum' else b_value} "
                            f"{'[PK, ' if a_value == 'ID' else '['} "
                            f"Note: '{c_value}' "
                            f"{',Not Null' if IsNotNull else ''}"
                            f"{f', Default: {DefaultKeyResult}' if DefaultKey else ''}] \n"
                        )
                        output_line = output_line.replace('#', '//')
                        output_text += output_line
            
            output_text += "}\n\n"
        
        # Foreign key 처리
        if 'Define' in workbook.sheetnames:
            sheet = workbook['Define']
            base_filename = filename.replace('.xlsx', '').replace('DataDefine_', '')
            
            # 헤더 찾기
            header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
            foreignkey_col_index = None
            for idx, cell_value in enumerate(header_row):
                if cell_value and cell_value.lower() == "foreignkey":
                    foreignkey_col_index = idx
                    break
            
            if foreignkey_col_index is not None:
                for row in sheet.iter_rows(min_row=2, values_only=True):
                    if len(row) > foreignkey_col_index:
                        fk_value = row[foreignkey_col_index]
                        property_a = row[0]
                        if fk_value is not None and property_a is not None:
                            output_text += f"Ref: {base_filename}.{property_a} {'<'} {fk_value}\n"
        
        # Enum 데이터 수집
        enum_data = []
        if 'Enum' in workbook.sheetnames:
            sheet = workbook['Enum']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 4:
                    enum_type = row[0]
                    enum_value = row[1]
                    enum_description = row[4]
                    if enum_type and enum_value and enum_description:
                        enum_data.append((enum_type, enum_value, enum_description))
        
        # TableGroup 데이터 수집 (유료 버전)
        tablegroup_data = []
        if is_paid_version and 'TableGroup' in workbook.sheetnames:
            sheet = workbook['TableGroup']
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if len(row) > 1:
                    TableGroup = row[0]
                    TableName = row[1]
                    if TableGroup and TableName:
                        tablegroup_data.append((TableGroup, TableName))
        
        workbook.close()
        return output_text, enum_data, tablegroup_data
        
    except Exception as e:
        st.error(f"파일 처리 중 오류 발생 ({filename}): {str(e)}")
        return "", [], []

def process_excel_files(uploaded_files, is_paid_version):
    """여러 Excel 파일 처리"""
    if not uploaded_files:
        return "오류: 선택된 Excel 파일이 없습니다."
    
    output_text = ""
    enum_dict = {}
    tablegroup_dict = {}
    
    # 진행 상황 표시
    progress_bar = st.progress(0)
    status_text = st.empty()
    
    try:
        # 각 파일 처리
        total_files = len(uploaded_files)
        for index, file in enumerate(uploaded_files):
            status_text.text(f"처리 중: {file.name}")
            
            # 파일 크기 확인
            file_size = len(file.getvalue()) / (1024 * 1024)  # MB 단위
            if file_size > 1000:  # 1GB 이상
                st.warning(f"경고: {file.name}의 크기가 큽니다 ({file_size:.1f}MB). 처리에 시간이 걸릴 수 있습니다.")
            
            # 파일 처리
            text, enum_data, tablegroup_data = process_excel_file(file, is_paid_version)
            output_text += text
            
            # Enum 데이터 처리
            for enum_type, enum_value, enum_description in enum_data:
                if enum_type not in enum_dict:
                    enum_dict[enum_type] = []
                enum_dict[enum_type].append(f"{enum_value} [note: '{enum_description}']")
            
            # TableGroup 데이터 처리
            for group, name in tablegroup_data:
                if group not in tablegroup_dict:
                    tablegroup_dict[group] = []
                tablegroup_dict[group].append(name)
            
            # 진행률 업데이트
            progress_bar.progress((index + 1) / total_files)
        
        # Enum 데이터 출력
        for enum_type, values in enum_dict.items():
            output_text += f"enum {enum_type}\n{{\n"
            for value in values:
                output_text += f"{value}\n"
            output_text += "}\n\n"
        
        # TableGroup 출력 (유료 버전)
        if is_paid_version:
            for TableGroup, values in tablegroup_dict.items():
                output_text += f"TableGroup {TableGroup}\n{{\n"
                for value in values:
                    output_text += f"{value}\n"
                output_text += "}\n\n"
        
        status_text.text("처리 완료!")
        progress_bar.progress(100)
        
        return output_text
        
    except Exception as e:
        return f"오류 발생: {str(e)}"

def main():
    st.set_page_config(
        page_title="DataDefine ERD Exporter",
        layout="wide",
        initial_sidebar_state="expanded"
    )
    
    st.title("DataDefine ERD Exporter")
    st.write("Excel 파일들을 선택하여 ERD 텍스트를 생성하세요.")
    
    # 사이드바에 컨트롤 배치
    with st.sidebar:
        st.header("설정")
        
        # 유료 버전 체크박스
        is_paid_version = st.checkbox("유료버전(TableGroup, HeaderColor)", value=True)
        
        # 파일 업로더
        uploaded_files = st.file_uploader(
            "Excel 파일 선택 (여러 파일 선택 가능)",
            type=['xlsx'],
            accept_multiple_files=True,
            help="DataDefine Excel 파일들을 선택하세요."
        )
        
        # 선택된 파일 목록 표시
        if uploaded_files:
            st.write(f"선택된 파일 수: {len(uploaded_files)}")
            file_sizes = [len(file.getvalue())/(1024*1024) for file in uploaded_files]
            total_size = sum(file_sizes)
            st.write(f"총 크기: {total_size:.1f}MB")
            
            with st.expander("선택된 파일 목록"):
                for file, size in zip(uploaded_files, file_sizes):
                    st.write(f"- {file.name} ({size:.1f}MB)")
        
        # dbdiagram.io 링크
        st.markdown("---")
        st.markdown("[dbdiagram.io에서 열기](https://dbdiagram.io/d/)")
    
    # 메인 영역
    if uploaded_files:
        if st.button("ERD 텍스트 생성", key="generate_button"):
            with st.spinner('ERD 텍스트 생성 중...'):
                result = process_excel_files(uploaded_files, is_paid_version)
                
                if result:
                    # 결과 표시
                    st.text_area("생성된 ERD 텍스트", result, height=500, key="result_area")
                    
                    # 결과를 다운로드할 수 있는 버튼 추가
                    st.download_button(
                        label="텍스트 파일로 다운로드",
                        data=result,
                        file_name="ERD_text.txt",
                        mime="text/plain",
                    )
                    
                    # 복사 버튼 (JavaScript)
                    st.markdown("""
                        <button onclick="
                            navigator.clipboard.writeText(document.querySelector('textarea').value)
                            .then(() => alert('텍스트가 클립보드에 복사되었습니다!'))
                            .catch(err => alert('복사 실패: ' + err))
                        ">클립보드에 복사</button>
                    """, unsafe_allow_html=True)

if __name__ == "__main__":
    main()